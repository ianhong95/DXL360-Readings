from readingUI import readingUI, internalController
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem

import sys
import serial

import openpyxl as xl

from SerialObjClass import SerialObj as cereal


def takeReading(device):
    rawBytes = device.serialRead()
    reading = device.translate()
    return reading


def recordReading(mainGUI, device):
    reading = takeReading(device)

    floatReading = str(reading[2])

    if mainGUI.readingTbl.currentRow() == -1:
        mainGUI.readingTbl.setCurrentCell(0, 1)
        mainGUI.readingTbl.setItem(mainGUI.readingTbl.currentRow(), 1, QTableWidgetItem(floatReading))
        mainGUI.readingTbl.setCurrentCell(mainGUI.readingTbl.currentRow() + 1, 1)

    elif mainGUI.readingTbl.currentRow() != -1 and mainGUI.readingTbl.currentRow() < mainGUI.readingTbl.rowCount() - 1:
        mainGUI.readingTbl.setItem(mainGUI.readingTbl.currentRow(), 1, QTableWidgetItem(floatReading))
        mainGUI.readingTbl.setCurrentCell(mainGUI.readingTbl.currentRow() + 1, 1)

    elif mainGUI.readingTbl.currentRow() != -1 and mainGUI.readingTbl.currentRow() == mainGUI.readingTbl.rowCount() - 1:
        mainGUI.readingTbl.setItem(mainGUI.readingTbl.currentRow(), 1, QTableWidgetItem(floatReading))


class xlWorkbook:
    def __init__(self, filePath, fileName):
        self.filePath = filePath
        self.fileName = fileName
        self.xlWb = self.initWorkbook()
    

    def initWorkbook(self):
        self.xlWb = xl.load_workbook(self.filePath + '\\' + self.fileName)
        return self.xlWb
        

class externalController:
    def __init__(self, mainGUI):
        self.mainGUI = mainGUI

        self.connectExtSigs()
    
    
    def initCOMPort(self):
        if self.mainGUI.startButton.text() == 'Stop':
            self.COMPort = 'COM' + self.mainGUI.lineEdits['edCOMPort'].text()
            try:
                self.serialDevice = cereal(self.COMPort, 9600)
            except:
                print('Try again!')


    # ---- Excel handling functions ----- #
    def xlWb(self):
        self.xlPath = self.mainGUI.lineEdits['edFilePath'].text()
        self.xlFile = self.mainGUI.lineEdits['edFileName'].text() + '.xlsx'

        # Initialize excel workbook object
        wbObject = xlWorkbook(self.xlPath, self.xlFile)
        self.wb = wbObject.xlWb
        self.source = self.wb['Template']


    def xlWrite(self):
        armID = self.mainGUI.lineEdits['edArmID'].text()
        self.workSheet = self.wb.copy_worksheet(self.source)
        self.workSheet.title = armID

        dateCell = self.workSheet['C2']
        armIDCell = self.workSheet['C3']
        armDescCell = self.workSheet['C4']

        dateCell.value = self.mainGUI.lineEdits['edDate'].text()
        armIDCell.value = self.mainGUI.lineEdits['edArmID'].text()
        armDescCell.value = self.mainGUI.lineEdits['edArmDesc'].text()

        tblRowCount = self.mainGUI.count

        for row in range(tblRowCount):
            try:
                positionValue = self.mainGUI.readingTbl.item(row, 0).text()
                dataValue = self.mainGUI.readingTbl.item(row, 1).text()

                positionCell = self.workSheet['B' + str(row + 6)]
                dataCell = self.workSheet['C' + str(row + 6)]

                dataCell.value = float(dataValue)
                positionCell.value = positionValue
                
            except:
                print('no data!')

        self.wb.save(self.xlPath + '\\' + self.xlFile)
        self.mainGUI.statusLabel.setText('Data exported!')
        print('Data exported!')


    def connectExtSigs(self):
        self.mainGUI.readButton.clicked.connect(lambda: recordReading(self.mainGUI, self.serialDevice))
        self.mainGUI.startButton.clicked.connect(self.initCOMPort)
        self.mainGUI.startButton.clicked.connect(self.xlWb)
        self.mainGUI.exportButton.clicked.connect(self.xlWrite)


def main():
    DXL360_App = QtWidgets.QApplication(sys.argv)

    mainGUI = readingUI()

    intController = internalController(mainGUI)
    extController = externalController(mainGUI)

    mainGUI.show()

    sys.exit(DXL360_App.exec())


if __name__ == '__main__':
    main()